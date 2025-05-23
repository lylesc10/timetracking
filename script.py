import pandas as pd
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
import tkinter as tk
from tkinter import messagebox, filedialog
from PIL import Image, ImageTk
import matplotlib.pyplot as plt

#to do : 
# - docker container 
# - github page 

def create_gui():
    file_data = []
    current_index = [0]

    def update_display():
        if not file_data:
            listbox.delete(0, tk.END)
            chart_label.config(image='', text="No data loaded")
            return
        data = file_data[current_index[0]]
        df = data['df']
        filename = data['filename']
        # Update listbox
        listbox.delete(0, tk.END)
        listbox.insert(tk.END, f"File: {os.path.basename(filename)}")
        listbox.insert(tk.END, "Comment Summary:")
        for index, row in df.iterrows():
            listbox.insert(tk.END, f"  - {row['Comment']}: {row['Total Hours']}")
        # Update chart
        display_chart(df)

    def prev_file():
        if file_data:
            current_index[0] = (current_index[0] - 1) % len(file_data)
            update_display()

    def next_file():
        if file_data:
            current_index[0] = (current_index[0] + 1) % len(file_data)
            update_display()

    def process_files():
        # Open a file dialog to select Excel files
        files = filedialog.askopenfilenames(
            title="Select Excel Files",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not files:
            return  # If no files are selected, do nothing

        file_data.clear()  # Clear previous data
        # Clear the listbox before displaying new column labels
        listbox.delete(0, tk.END)

        df = pd.DataFrame() 
        for file in files:
            if file.lower().endswith(('.xlsx', '.xls')):
                try:
                    # Process the file
                    format_training_time(file)

                    # Load the second sheet from the processed file
                    base_name = os.path.splitext(file)[0]
                    output_file_path = f"{base_name}_sorted.xlsx"
                    df = pd.read_excel(output_file_path, sheet_name="Comment Summary")
                    file_data.append({'df': df, 'filename': file})
                    # Display the content of the second sheet in the listbox
                    listbox.insert(tk.END, f"File: {os.path.basename(output_file_path)}")
                    listbox.insert(tk.END, "Comment Summary:")
                    for index, row in df.iterrows():
                        listbox.insert(tk.END, f"  - {row['Comment']}: {row['Total Hours']}")

                    messagebox.showinfo("Success", f"Processed and displayed file: {file}")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to process file: {file}\n{str(e)}")
            else:
                messagebox.showwarning("Invalid File", f"Skipped non-Excel file: {file}")
        current_index[0] = 0
        update_display()
    # Create the main window
    root = tk.Tk()
    root.title("Excel File Processor")
    root.geometry("1080x720")

    # Change the background color of the main window
    #root.configure(bg="#f0f0f0")

    # Add instructions label
    label = tk.Label(root, text="Click the button below to select Excel files to process", font=("Arial", 14))
    label.pack(pady=10)

    # Add a button to open the file dialog
    button = tk.Button(root, text="Select Files", command=process_files, font=("Arial", 12))
    button.pack(pady=10)

    global chart_label
    chart_label = tk.Label(root, text="Bar Chart of Total Hours per Unique Comment")
    chart_label.pack(pady=10)

    nav_frame = tk.Frame(root)
    nav_frame.pack(pady=5)
    prev_btn = tk.Button(nav_frame, text="Previous File", command=prev_file)
    prev_btn.pack(side=tk.LEFT, padx=5)
    next_btn = tk.Button(nav_frame, text="Next File", command=next_file)
    next_btn.pack(side=tk.LEFT, padx=5)

    # Add a listbox to display the content of the second sheet
    listbox_label = tk.Label(root, text="Content from the 'Comment Summary' Sheet:", font=("Arial", 12))
    listbox_label.pack(pady=5)
    listbox = tk.Listbox(root, width=100, height=30, bg="#ffffff", fg="#000000")
    listbox.pack(pady=10)

    # Run the GUI event loop
    root.mainloop()



def display_chart(df):
        # Generate a bar chart using matplotlib
        plt.figure(figsize=(8, 6))
        plt.bar(df['Comment'], df['Total Hours'], color='skyblue')
        plt.title('Total Hours per Unique Comment')
        plt.xlabel('Comment')
        plt.ylabel('Total Hours')
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        plt.savefig("chart.png")
        plt.close()

        # Load and display the image in the GUI
        img = Image.open("chart.png")
        img = img.resize((600, 400), Image.Resampling.LANCZOS)
        img_tk = ImageTk.PhotoImage(img)
        chart_label.config(image=img_tk)
        chart_label.image = img_tk

def format_training_time(input_file):
    # Get file type
    _, file_type = os.path.splitext(input_file)

    if file_type.lower() not in ['.xlsx', '.xls']:
        print(f"Error: The file is not an excel file")
        return False

    # Load the input file
    df = pd.read_excel(input_file, sheet_name="Sheet 1")

    # Filter out Interns
    df = df[df["Department"] != 'Interns']

    # Fill Week End Date
    if "Week End Date" in df.columns:
        df["Week End Date"] = df["Week End Date"].ffill()

    # Remove the "Total" row if present
    df = df[df['Week End Date'] != 'Grand Total']

    # Select and rename relevant columns
    df = df[["Supervisor Name", "Week End Date", "Employee Name", "Task Number/Name", "Comments", "Person Id", "Hours"]]

    # Convert "Hours" column to numeric
    df["Hours"] = pd.to_numeric(df["Hours"], errors='coerce')

    # Compute total hours per employee
    df["Total hours"] = df.groupby("Employee Name")["Hours"].transform("sum")

    df = df.sort_values(by=["Supervisor Name", "Employee Name"])

    # Sort the data
    df = df.sort_values(by=["Supervisor Name", "Employee Name"])

    # Add a row for each supervisor with the sum of hours for their employees
    supervisor_totals = df.groupby("Supervisor Name")["Hours"].sum().reset_index()
    supervisor_totals["Employee Name"] = "Total for Supervisor"
    supervisor_totals["Task Number/Name"] = ""
    supervisor_totals["Comments"] = ""
    supervisor_totals["Person Id"] = ""
    supervisor_totals["Week End Date"] = ""

    # Append the totals to the original DataFrame
    df = pd.concat([df, supervisor_totals], ignore_index=True)

    # Sort again to ensure totals appear after each supervisor's employees
    df = df.sort_values(by=["Supervisor Name", "Employee Name"], key=lambda col: col.map(lambda x: (x == "Total for Supervisor", x)))

    # Save to a new Excel file
    base_name = os.path.splitext(input_file)[0]
    output_file_path = f"{base_name}_sorted.xlsx"

    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Sorted Data")

        # Save comments and hours to another sheet
        comment_hours_df = df.groupby("Comments", dropna=False)["Hours"].sum().reset_index()
        comment_hours_df.columns = ["Comment", "Total Hours"]

        #this line is what takes out the no comments on the second sheet
        comment_hours_df = comment_hours_df[comment_hours_df["Comment"] != "(No Comments Entered)"]
        comment_hours_df = comment_hours_df[comment_hours_df["Comment"] != ""]
        comment_hours_df = comment_hours_df.sort_values(by="Total Hours", ascending=False)
        comment_hours_df.to_excel(writer, index=False, sheet_name="Comment Summary")
        display_chart(comment_hours_df)

    # Reopen with openpyxl to style headers and add image
    wb = load_workbook(output_file_path)
    ws = wb["Sorted Data"]

    # Style headers green
    green_fill = PatternFill(start_color="459452", end_color="459452", fill_type="solid")
    for cell in ws[1]:
        cell.fill = green_fill

    # Merge cells in the Supervisor Name column that have the same value
    supervisor_col = 1  # Assuming Supervisor Name is column A (1-indexed)
    start_row = 2       # Data starts from row 2 (row 1 is header)
    end_row = ws.max_row

    current_supervisor = ws.cell(row=start_row, column=supervisor_col).value
    merge_start = start_row

    for row in range(start_row + 1, end_row + 1):
        cell_value = ws.cell(row=row, column=supervisor_col).value
        if cell_value != current_supervisor:
            if row - 1 > merge_start:
                ws.merge_cells(start_row=merge_start, start_column=supervisor_col,
                            end_row=row - 1, end_column=supervisor_col)
                ws.cell(row=merge_start, column=supervisor_col).alignment = Alignment(horizontal="center", vertical="center")
            current_supervisor = cell_value
            merge_start = row
    # Merge the last group
    if end_row > merge_start:
        ws.merge_cells(start_row=merge_start, start_column=supervisor_col,
                    end_row=end_row, end_column=supervisor_col)
        ws.cell(row=merge_start, column=supervisor_col).alignment = Alignment(horizontal="center", vertical="center")

    # Insert Excel-generated bar chart in Comment Summary
    comment_ws = wb["Comment Summary"]
    max_row = comment_ws.max_row

    chart = BarChart()
    chart.title = "Total Hours per Unique Comment"
    chart.x_axis.title = "Comment"
    chart.x_axis.titleOverlay = False
    chart.y_axis.title = "Total Hours"
    chart.y_axis.titleOverlay = False
    chart.width = 40
    chart.height = 20
    chart.style = 3

    #graph with all comments
    data = Reference(comment_ws, min_col=2, min_row=1, max_row=max_row)
    categories = Reference(comment_ws, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    comment_ws.add_chart(chart, "D40")

    chart = BarChart()
    chart.title = "Total Hours per Unique Comment"
    chart.x_axis.title = "Comment"
    chart.x_axis.titleOverlay = True
    chart.y_axis.title = "Total Hours"
    chart.y_axis.titleOverlay = True
    chart.width = 40
    chart.height = 20
    chart.style = 3

    #take the top 10 comments, max row is the important bit 
    data = Reference(comment_ws, min_col=2, min_row=1, max_row=11)
    categories = Reference(comment_ws, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    comment_ws.add_chart(chart, "D1")

    #display_chart(comment_hours_df)
    # Save final workbook
    wb.save(output_file_path)
    print(f"Sorted the excel file and saved it as '{output_file_path}'")

if __name__ == "__main__":
    if len(sys.argv) == 1:
        create_gui()
    elif len(sys.argv) == 2:
        input_file = sys.argv[1]
        format_training_time(input_file)
    else:
        print("Usage: python script.py [<input_excel_file>]")
